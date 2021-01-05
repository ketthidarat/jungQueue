from django.core.management.base import BaseCommand, CommandError
from api.models import *
from openpyxl import load_workbook 


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print('กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A2'].value)
        print(f'count = {count}')
        #row4 = [ ws[f'{c}4'].value for c in 'ABCDEFG' ]
        #print( row4 )
        #column_names = ['id', 'rice_type']
        data = []  
        for i in range(count): # 0,1,2,3
            print(f'i = {i}')
            sheet_values = [ ws[f'{chr(65+j)}{4+i}'].value for j in range(len(column_names)) ]
            data.append( dict( (k,v) for k,v in zip(column_names, sheet_values)) )

        return data

    def handle(self, *args, **options):
        # pass
        from openpyxl import load_workbook
        filename = "xlsx/แผนการดำเนินโครงงานที่ปรึกษา senior project 62 - เกษ.xlsx"
        wb = load_workbook(filename, data_only=True)
        #sheets = [ 'Rice_type', 'Work_status', 'Money_status', 'Work' ]

        print('กำลัง load ... Rice_type')
        for d in self.load(wb, 'Rice_type', ['id', 'rice_type']):
            print(d)
            Rice_type(**d).save()

        print('กำลัง load ... Work_status')
        for d in self.load(wb, 'Work_status', ['id', 'work_status']):
            Work_status(**d).save()

        print('กำลัง load ... Money_status')
        for d in self.load(wb, 'Money_status', ['id', 'money_status']):
            Money_status(**d).save()

        print('กำลัง load ... Farmer')
        for d in self.load(wb, 'Farmer', ['id', 'farmer_name','phone', 'address', 'email', 'username', 'password', 'image']):
            print(d)
            Farmer(**d).save()

        print('กำลัง load ... Owner')
        for d in self.load(wb, 'Owner', ['id', 'owner_name','phone', 'address', 'email', 'username', 'password', 'image']):
            print(d)
            Owner(**d).save()

        print('กำลัง load ... Tractor_status')
        for d in self.load(wb, 'Tractor_status', ['id', 'tractor_status']):
            print(d)
            Tractor_status(**d).save()
        
        print('กำลัง load ... Tractor')
        for d in self.load(wb, 'Tractor', ['id', 'tractor', 'tractor_status']):
            print(d)
            Tractor(**d).save()


        # print('กำลัง load ... Work')
        # for d in self.load(wb, 'Work', ['id', 'lat', 'lng', 'date_start', 'date_end', 'area', 'rice_type', 'other', 'RepairTime', 'Harverstime', 'money', 'money_status', 'work_status', 'tractor', 'tractor_status']):
        #     Work(**d).save()

        ''' Bug Fixed
        rice_type = Rice_type.objects.get(pk=d['rice_type'])
        d.pop('rice_type')
        w = Work(**d)
        w.rice_type = rice_type
        ....
        w.save()
        '''

"""
        ws = wb['Work']
        count = int(ws['A2'].value)
        print(f'count = {count}')
        #row4 = [ ws[f'{c}4'].value for c in 'ABCDEFG' ]
        #print( row4 )
        for i in range(count): # 0,1,2,3
            #print(f'i = {i}')
            w = Work(**{
                    'id': int(ws[f'A{4+i}'].value),
                    'lat': float(ws[f'B{4+i}'].value),
                    'lng': float(ws[f'C{4+i}'].value),
                    'date': str(ws[f'D{4+i}'].value),
                    'area': float(ws[f'E{4+i}'].value),
                    'rice_type': Rice_type.objects.get(pk=int(ws[f'F{4+i}'].value)),
                    'other': str(ws[f'G{4+i}'].value),
                    'RepairTime': str(ws[f'H{4+i}'].value),
                    'Harverstime': str(ws[f'I{4+i}'].value),
                    'money': int(ws[f'J{4+i}'].value),
                    'moneyStatus': Money_status.objects.get(pk=int(ws[f'K{4+i}'].value)),
                    'work_status ': Work_status.objects.get(pk=int(ws[f'L{4+i}'].value)),
                })
            w.save()
        """

     
            # f = Farmer(**{
            #         'id': int(ws[f'A{4+i}'].value),
            #         'farmer_name': str(ws[f'B{4+i}'].value),
            #         'phone': str(ws[f'C{4+i}'].value),
            #         'address': str(ws[f'D{4+i}'].value),
            #         'email': str(ws[f'E{4+i}'].value),
            #         'username': str(ws[f'F{4+i}'].value),
            #         'password': str(ws[f'G{4+i}'].value),
                    
            #     })
            # f.save()

    
            # r = Owner(**{
            #     #    'id': int(ws[f'A{4+i}'].value),
            #         'owner_name': str(ws[f'B{4+i}'].value),
            #         'phone': str(ws[f'C{4+i}'].value),
            #         'address': str(ws[f'D{4+i}'].value),
            #         'email': str(ws[f'E{4+i}'].value),
            #         'username': str(ws[f'F{4+i}'].value),
            #         'password': str(ws[f'G{4+i}'].value),
                    
            #     })
            # r.save()
        #     w = Work(**{
        #             'id': int(ws[f'A{4+i}'].value),
        #             'lat': float(ws[f'B{4+i}'].value),
        #             'lng': float(ws[f'C{4+i}'].value),
        #             'date': str(ws[f'D{4+i}'].value),
        #             'area': float(ws[f'E{4+i}'].value),
        #             # 'rice_type': str(ws[f'F{4+i}'].value),
        #             'other': str(ws[f'G{4+i}'].value),
        #             'RepairTime': str(ws[f'H{4+i}'].value),
        #             'Harverstime': str(ws[f'I{4+i}'].value),
        #             'money': str(ws[f'J{4+i}'].value),
        #             # 'moneyStatus': str(ws[f'K{4+i}'].value),
        #             # 'work_status ': str(ws[f'L{4+i}'].value),
                    
        #         })
        # w.save()

    
            # t = Rice_type(**{
            #         # 'id': int(ws[f'A{4+i}'].value),
            #         'rice_type': str(ws[f'B{4+i}'].value),
                   
                    
                    
            #     })
            # t.save()

            # t = Tractor(**{
            #         # 'id': int(ws[f'A{4+i}'].value),
            #         'tractor': str(ws[f'B{4+i}'].value),
                   
                    
                    
            #     })
            # t.save()


            # t = Work_status(**{
            #         # 'id': int(ws[f'A{4+i}'].value),
            #         'work_status': str(ws[f'B{4+i}'].value),
                   
                    
                    
            #     })
            # t.save()

            # t = Money_status(**{
            #         # 'id': int(ws[f'A{4+i}'].value),
            #         'moneyStatus': str(ws[f'B{4+i}'].value),
                   
                    
                    
            #     })
            # t.save()

